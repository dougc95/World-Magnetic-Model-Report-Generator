import WMMv2
import calendar
from datetime import datetime
import openpyxl


class WMM:
    def __init__(self,latitude,longitude,altitude,year):
        self.latitude = latitude
        self.longitude = longitude
        self.altitude = altitude
        self.year = year
        #self.file = file

    def callWMM(self):
        self.dec=WMMv2.getDeclination(self.latitude,self.longitude,self.year,self.altitude)
        self.dip=WMMv2.getDipAngle(self.latitude,self.longitude,self.year,self.altitude)
        self.by=WMMv2.getEastIntensity(self.latitude,self.longitude,self.year,self.altitude)
        self.bh=WMMv2.getHorizontalIntensity(self.latitude,self.longitude,self.year,self.altitude)
        self.ti=WMMv2.getIntensity(self.latitude,self.longitude,self.year,self.altitude)
        self.bx=WMMv2.getNorthIntensity(self.latitude,self.longitude,self.year,self.altitude)
        self.bz=WMMv2.getVerticalIntensity(self.latitude,self.longitude,self.year,self.altitude)
        #return dec, dip, by, bh, ti, bx, bz
    
    def decimalYear(self,year):
        date_object = datetime.strptime(self.year, '%Y-%m-%d').date()
        daysInYear = 0
        if (calendar.isleap(date_object.year)):
            daysInYear=366
        else:
            daysInYear=365
        #self.year =date_object.year+(date_object.timetuple().tm_yday/daysInYear) 
        #print(self.year)
        return date_object.year+(date_object.timetuple().tm_yday/daysInYear)#Date to day of year

    def xlsGenerationTopics(self):
        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.title = "WMM"
        print(hoja.title)
        #Topics
        hoja.cell(row=1, column=1, value = 'Date')
        hoja.cell(row=1, column=2, value = 'Total Field')
        hoja.cell(row=1, column=3, value = 'Horizontal')
        hoja.cell(row=1, column=4, value = 'North')
        hoja.cell(row=1, column=5, value = 'East')
        hoja.cell(row=1, column=6, value = 'Vertical')
        hoja.cell(row=1, column=7, value = 'Declination')
        hoja.cell(row=1, column=8, value = 'Inclination')
        #Variation Change/Year
        hoja.cell(row=1, column=10, value = '\u0394 Total Field')
        hoja.cell(row=1, column=11, value = 'Delta Horizontal')
        hoja.cell(row=1, column=12, value = 'Delta North')
        hoja.cell(row=1, column=13, value = 'Delta East')
        hoja.cell(row=1, column=14, value = 'Delta Vertical')
        hoja.cell(row=1, column=15, value = 'Delta Declination')
        hoja.cell(row=1, column=16, value = 'Delta Inclination')

        wb.save("prueba.xlsx")

        pass

    


# World = WMM(lat,lon,alt,yr)
# World.decimalYear()
# print(World.year)
# World.callWMM()
#World.xlsGeneration()



'''
print(World.latitude)
print(World.longitude)
print(World.altitude)
print(World.year)
''' 
#print(World.dec)